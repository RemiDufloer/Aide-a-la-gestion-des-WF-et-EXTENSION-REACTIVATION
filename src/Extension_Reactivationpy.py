# -*- coding: utf-8 -*-
"""
Created on Tue Feb 21 16:34:22 2023

@author: rechacalos
"""

# Extension Reactivation IHM

import sys
import openpyxl
import choix_utilisateur as cu
import WF_i00 as wf
import time

# start  of the content
#regarder si c'est chez ced ou pas donc pas forcement 6600
# donc prendre article cu .xlsx
def give_number_article_by_designation_article():
    wb = openpyxl.load_workbook(filename='../ressources/articles CU.XLSX', read_only=True)
    ws = wb.active
    res_number = []
    res_designation = []
    res_div = []
    designation_article = cu.demande_a_utilisateur_string("désignation article :")
    compteur =0
    for i in ws.rows :
        compteur +=1
        if compteur == 1:
            continue     
        if (designation_article.upper() in str(i[1].value) ):
            res_number.append(i[0].value)
            res_designation.append((i[1].value))
            res_div.append((i[2].value))
    if res_number == []:    
        print("cette article n'est pas référencé dans articles CU.XLSX")
    wb.close()
    return res_designation, res_number,res_div


def give_division_article_by_number(number):
    wb = openpyxl.load_workbook(filename='../ressources/articles CU.XLSX', read_only=True)
    ws = wb.active
    division = -1
    compteur =0
    for i in ws.rows :
        compteur +=1
        if compteur == 1:
            continue    
        if (number == i[0].value ):
            division = str(i[2].value)
    if division == -1:    
        print("cette article n'est pas référencé ")
    elif division == "6600" :
        print("cette article est référencé en 6600 ")
    else :
        print("cette article n'est pas référencé en 6600 mais en " + str(division))
    wb.close()
    return division

def liste_diffférent_article():
    while True :
        res_designation, res_number, res_div = give_number_article_by_designation_article()
        i = 0
        while i < len(res_designation):
            if res_designation[i].startswith("DBL"):
                res_designation.remove(res_designation[i])
                res_number.remove(res_number[i])
                res_div.remove(res_div[i])
                i-=1
            elif res_designation[i].startswith("OBS"):
                res_designation.remove(res_designation[i])
                res_number.remove(res_number[i])
                res_div.remove(res_div[i])
                i-=1
            elif res_designation[i].startswith("HRM"):
                res_designation.insert(0, res_designation[i])
                res_number.insert(0, res_number[i])
                res_div.insert(0, res_div[i])
                res_designation.remove(res_designation[i+1])
                res_number.remove(res_number[i+1])
                res_div.remove(res_div[i+1])
            i+=1
        i = 0
        for i in range(len(res_designation)):
            print(res_designation[i]," | ",res_number[i],"\n")
        
        result_number = cu.demande_a_utilisateur_string("quelle article voulait vous parmis la liste ? si aucun taper -1 :")
        
        if int(result_number) != -1 :
            while True:        
                if int(result_number) != -1 :
                    try:
                        index = res_number.index((result_number))
                        result_designation= res_designation[index]
                        div = res_div[index]
                
                        break
                
                    except ValueError:
                
                        print("Oops!  That was no valid number.  Try again...")
                        result_number = cu.demande_a_utilisateur_string("quelle article voulait vous parmis la liste ? si aucun taper -1 :")
                else:
                    break
            if int(result_number) != -1 :
                break
    result_number = check_division(result_number,div)
      
    return result_designation,result_number


def check_division(result_number,div):
    if div == "6600" :
        return -1
    else :
        print("cette article n'est pas référencé en 6600 mais en " + str(div))
    return result_number

    

def secteur_demandeur():
    choix = cu.demande_a_utilisateur("Quelle est le secteur demandeur ?  \n 1 | APREG \n 2 | CHAUD \n 3 | ERA \n 4 | MECA \n 5 | Autre - Magasin ",1,5)
    res = 0
    if choix == 1 :
        res = "APREG"
    elif choix == 2 :
        res = "CHAUD"
    elif choix == 3 :
        res = "ERA"
    elif choix == 4 :
        res = "MECA"
    elif choix == 5 :
        res = "Autre - Magasin"
    return res 

def Extension_ou_Reactivation():
    choix = cu.demande_a_utilisateur("Extension ou Réactivation ?  \n 1 | Réactivation \n 2 | Extension \n  ",1,2)
    res = 0
    if choix == 1 :
        res = 'REACTIVATION'
    elif choix == 2 :
        res = 'EXTENSION'
    return res 

def Emplacement():
    return cu.demande_a_utilisateur_string("Quelle Emplacement dans le magasin ?")


def exec_of_main():
    wb = openpyxl.load_workbook('../wf.xlsx')
    result_designation,result_number = liste_diffférent_article()
    if result_number == -1 :
        print("il n'y a donc rien a faire, vous pouvez fermer ce programme, ou il se fermera automatiquement dans 3 seconde")
        time.sleep(3)
        return
    secteurDemandeur = secteur_demandeur()
    Type_Demande = Extension_ou_Reactivation()
    y = wf.give_first_empty_line(wb[Type_Demande])
    numero_equipements = wf.search_equipement_by_article(result_number)
    emplacement = Emplacement()
    if numero_equipements == []:
        wf.extension_reactivation(wb[Type_Demande], y, "Chauvet", secteurDemandeur, result_number, result_designation, emplacement, "?", "", "", "", "", "")
    for numero_equipement in numero_equipements :
        denomination , Designation_equipement, Numéro_poste_technique, poste_technique = wf.remplissage_extension_reactivation_equipement(numero_equipement)
        wf.extension_reactivation(wb[Type_Demande], y, "Chauvet", secteurDemandeur, result_number, result_designation, emplacement, "?", numero_equipement, denomination, Designation_equipement, poste_technique, Numéro_poste_technique)
    wb.save("wf.xlsx")

if __name__ == '__main__':
   exec_of_main()