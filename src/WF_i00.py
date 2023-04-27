import openpyxl


def init_WF():
    workbook = openpyxl.Workbook()
    workbook.active.title = 'EXTENSION'
    workbook.create_sheet('REACTIVATION')
    workbook.create_sheet('CREATION ARTICLE')
    workbook.create_sheet('Mise à jour tarifaire')
    workbook.save('../wf.xlsx')
    return workbook

def extension_reactivation_init(sheet):
    sheet.cell(1, 2).value = 'Nom'   
    sheet.cell(1, 3).value = 'Secteur demandeur '
    sheet.cell(1, 4).value = "Origine Demandeur"
    sheet.cell(1, 5).value = "Type Demande "
    sheet.cell(1, 6).value = "Gestion du stock"
    sheet.cell(1, 7).value = "Groupe marchandise"
    sheet.cell(1, 8).value = 'code article  '
    sheet.cell(1, 9).value = 'Désignation article '
    sheet.cell(1, 10).value = 'Proposition Désignation'
    sheet.cell(1, 11).value = 'texte commande achat '
    sheet.cell(1, 12).value = 'T. Plan.'
    sheet.cell(1, 13).value = 'Point de commande'
    sheet.cell(1, 14).value = 'Type article'
    sheet.cell(1, 15).value = 'Justification'
    sheet.cell(1, 16).value = 'Emplacement'   
    sheet.cell(1, 17).value = 'Gestion magasin '
    sheet.cell(1, 18).value = 'Numéro d’équipement  '
    sheet.cell(1, 19).value = 'Dénomination '
    sheet.cell(1, 20).value = 'Désignation équipement  '
    sheet.cell(1, 21).value = 'Poste technique'
    sheet.cell(1, 22).value = 'Numéro poste technique'
    
def extension_reactivation(sheet,y,Nom,Secteur_demandeur,numero_article,designation_article,Emplacement,Gestion_magasin,numero_equipement,denomination,Designation_equipement,poste_technique,Numéro_poste_technique):
    sheet.cell(y, 2).value = Nom #'Nom'   
    sheet.cell(y, 3).value = Secteur_demandeur #'Secteur demandeur '
    sheet.cell(y, 8).value = numero_article #'Numéro d’article  '
    sheet.cell(y, 9).value = designation_article #'Désignation article '
    sheet.cell(y, 16).value = Emplacement #'Emplacement'   
    sheet.cell(y, 17).value = Gestion_magasin #'Gestion magasin '
    sheet.cell(y, 12).value = 'PD'
    sheet.cell(y, 18).value = numero_equipement #'Numéro d’équipement  '
    sheet.cell(y, 19).value = denomination #'Dénomination '
    sheet.cell(y, 20).value = Designation_equipement #'Désignation équipement  '
    sheet.cell(y, 21).value = poste_technique #'Poste technique'
    sheet.cell(y, 22).value = Numéro_poste_technique #'Numéro poste technique'


def give_first_empty_line(sheet):
    return sheet.max_row + 1 

def write_designation_article(designation_article,sheet):
    sheet.cell((give_first_empty_line(sheet)), 4).value = designation_article


def search_equipement_by_article(numero_article):
    stream_TTs = open('../ressources/ARBORESCENCE.txt')
    lines = stream_TTs.readlines()
    stream_TTs.close()
    res = True
    i = 0
    j = 0
    equipements = []
    while (res and i<len(lines)):
        info_ligne = info_tts_ligne(lines[i])
        if info_ligne == []:
            j = i - 1
            i += 1
            continue
        if info_ligne[0]==str(numero_article):
            equipements.append((info_tts_ligne(lines[i])[0]))
        i+=1
    if equipements == [] :
        print("aucun equipement rattacher à cette article")
    else :
        print(equipements)
    return equipements

def remplissage_extension_reactivation_equipement(numero_equipement):
    stream_TTs = open('../ressources/ARBORESCENCE.txt')
    stream_TTs.readline()
    stream_TTs.readline()
    stream_TTs.readline()
    stream_TTs.readline()
    stream_TTs.readline()
    line = stream_TTs.readline()
    Dénomination = 0
    Désignation_équipement = 0
    Poste_technique = 0
    Numéro_poste_technique = 0
    i = 0
    res = False
    while line != '':
        info_ligne = info_tts_ligne(line)
        Numéro_poste_technique,Poste_technique = check_if_is_poste_tech(Numéro_poste_technique,Poste_technique,info_ligne)
        for j in info_ligne :
            if (j.lstrip() == str(numero_equipement)):
                Dénomination,Désignation_équipement=collect_info(Dénomination,Désignation_équipement,info_ligne)
                res = True
        line = ''
        if not (res):
            line = stream_TTs.readline()
            i+=1
    print(i)
    print(Dénomination,Désignation_équipement)
    print(Numéro_poste_technique,Poste_technique)
    return Dénomination , Désignation_équipement, Numéro_poste_technique, Poste_technique
            
def info_tts_ligne(line):   
    info_ligne = line.rstrip().split("  ")
    info_ligne = ' '.join(info_ligne).split()
    delete = []
    for i in range(len(info_ligne)) :
        if (info_ligne[i][0]=='|'):
            delete.append(i)
    for x in reversed(delete):
        info_ligne.pop(x)
    return info_ligne

def check_if_is_poste_tech(Numéro_poste_technique,Poste_technique,info_ligne):
    if (len(info_ligne) != 0):
        if ('FLDUN' in info_ligne[0] ):
            Numéro_poste_technique=info_ligne[0]
            Poste_technique = " ".join([str(item) for item in info_ligne[1:]])
    return (Numéro_poste_technique,Poste_technique)
 
def collect_info(Dénomination,Désignation_équipement,info_ligne):
        if len(info_ligne)>1:
            Dénomination=info_ligne[0]
            Désignation_équipement = " ".join([str(item) for item in info_ligne[1:]])
        return (Dénomination,Désignation_équipement)
    
if __name__ == '__main__':
    wb=init_WF()
    extension_reactivation_init(wb['EXTENSION'])
    extension_reactivation_init(wb['REACTIVATION'])
    wb.save('../wf.xlsx')