import openpyxl


def init_WF():
    workbook = openpyxl.Workbook()
    workbook.active.title = 'EXTENSION'
    workbook.create_sheet('REACTIVATION')
    workbook.create_sheet('CREATION ARTICLE')
    workbook.create_sheet('Mise à jour tarifaire')
    workbook.save('wf.xlsx')
    return workbook

def extension_reactivation_init(sheet):
    sheet.cell(1, 1).value = 'Nom'   
    sheet.cell(1, 2).value = 'Secteur demandeur '
    sheet.cell(1, 3).value = 'Numéro d’article  '
    sheet.cell(1, 4).value = 'Désignation article '
    sheet.cell(1, 5).value = 'Emplacement'   
    sheet.cell(1, 6).value = 'Gestion magasin '
    sheet.cell(1, 7).value = 'Numéro d’équipement  '
    sheet.cell(1, 8).value = 'Dénomination '
    sheet.cell(1, 9).value = 'Désignation équipement  '
    sheet.cell(1,10).value = 'Poste technique'
    sheet.cell(1, 11).value = 'Numéro poste technique'

def give_designation_article(numero_article):
    wb = openpyxl.load_workbook(filename='article en 6600.XLSX', read_only=True)
    ws = wb.active
    compteur =0
    for i in ws.rows :
        compteur +=1
        if compteur == 1:
            continue
        if (int(i[0].value) == numero_article):
            wb.close()
            print(i[1].value)
            return i[1].value
    print("cette article n'est pas référencé dans article en 6600.XLSX")
    return None

def give_first_empty_line(sheet):
    return sheet.max_row + 1 

def write_designation_article(designation_article,sheet):
    sheet.cell((give_first_empty_line(sheet)), 4).value = designation_article

def remplissage_extension_reactivation_equipement(numero_equipement):
    stream_TTs = open('ARBORESCENCE.txt')
    stream_TTs.readline()
    stream_TTs.readline()
    stream_TTs.readline()
    stream_TTs.readline()
    stream_TTs.readline()
    line = stream_TTs.readline()
    Dénomination = 0
    Désignation_équipement = 0
    Poste_technique = 0
    Numéro_poste_technique = 0
    i = 0
    res = False
    while line != '':
        info_ligne = info_tts_ligne(line)
        Numéro_poste_technique,Poste_technique = check_if_is_poste_tech(Numéro_poste_technique,Poste_technique,info_ligne)
        for j in info_ligne :
            if (j.lstrip() == str(numero_equipement)):
                Dénomination,Désignation_équipement=collect_info(Dénomination,Désignation_équipement,info_ligne)
                res = True
        line = ''
        if not (res):
            line = stream_TTs.readline()
            i+=1
    print(i)
    print(Dénomination,Désignation_équipement)
    print(Numéro_poste_technique,Poste_technique)
    return i
            
def info_tts_ligne(line):   
    info_ligne = line.rstrip().split("  ")
    info_ligne = ' '.join(info_ligne).split()
    delete = []
    for i in range(len(info_ligne)) :
        if (info_ligne[i][0]=='|'):
            delete.append(i)
    for x in reversed(delete):
        info_ligne.pop(x)
    return info_ligne

def check_if_is_poste_tech(Numéro_poste_technique,Poste_technique,info_ligne):
    if (len(info_ligne) != 0):
        if ('FLDUN' in info_ligne[0] ):
            Numéro_poste_technique=info_ligne[0]
            Poste_technique = " ".join([str(item) for item in info_ligne[1:]])
    return (Numéro_poste_technique,Poste_technique)
 
def collect_info(Dénomination,Désignation_équipement,info_ligne):
        Dénomination=info_ligne[0]
        Désignation_équipement = " ".join([str(item) for item in info_ligne[1:]])
        return (Dénomination,Désignation_équipement)
    
if __name__ == '__main__':
    wb=init_WF()
    extension_reactivation_init(wb['EXTENSION'])
    extension_reactivation_init(wb['REACTIVATION'])
    wb.save('wf.xlsx')
    remplissage_extension_reactivation_equipement(10062509)
    remplissage_extension_reactivation_equipement(10073108)
    remplissage_extension_reactivation_equipement(10126620)
    remplissage_extension_reactivation_equipement(9161808)
    give_designation_article(9103652)
    write_designation_article(give_designation_article(9103652),(wb['EXTENSION']))
    wb.save('wf.xlsx')